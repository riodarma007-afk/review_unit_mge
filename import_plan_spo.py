import openpyxl
import pymysql
import sys
from datetime import datetime

# Database config
DB_HOST = '103.58.102.44'
DB_PORT = 3306
DB_USER = 'mge_planning'
DB_PASS = 'PlanningMGE2026'
DB_NAME = 'mge_planning_staging'
EXCEL_PATH = r'e:\project rio\Dashboard_optrack\db_paramater plan.xlsx'

def get_db_connection():
    return pymysql.connect(host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASS, database=DB_NAME)

def main():
    print(f"Loading Excel file {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    # 1. Build Mappings
    ws_map = wb['Standar Penamaan']
    ob_plan_map = {}
    coal_plan_map = {}
    for row in ws_map.iter_rows(min_row=3, max_row=ws_map.max_row, values_only=True):
        if len(row) > 2 and row[2] and row[1]:
            ob_plan_map[str(row[2]).strip()] = str(row[1]).strip()
        if len(row) > 8 and row[8] and row[7]:
            coal_plan_map[str(row[8]).strip()] = str(row[7]).strip()
            
    print(f"Loaded {len(ob_plan_map)} OB mappings and {len(coal_plan_map)} Coal mappings.")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # 2. Create Tables
        cur.execute('''
            CREATE TABLE IF NOT EXISTS plan_spo_coal (
                id BIGINT AUTO_INCREMENT PRIMARY KEY,
                date DATE,
                pit VARCHAR(100),
                losstime_name VARCHAR(100),
                plan_hours DECIMAL(10,4),
                INDEX idx_date_pit (date, pit),
                INDEX idx_losstime (losstime_name)
            )
        ''')
        
        cur.execute('''
            CREATE TABLE IF NOT EXISTS plan_spo_ob (
                id BIGINT AUTO_INCREMENT PRIMARY KEY,
                date DATE,
                owner VARCHAR(50),
                pit VARCHAR(100),
                losstime_name VARCHAR(100),
                plan_hours DECIMAL(10,4),
                INDEX idx_date_pit (date, pit),
                INDEX idx_losstime (losstime_name)
            )
        ''')
        
        # Clear existing data so we can rerun safely
        cur.execute('TRUNCATE TABLE plan_spo_coal')
        cur.execute('TRUNCATE TABLE plan_spo_ob')
        print("Tables created and truncated successfully.")
        
        # 3. Import Coal SPO
        ws_coal = wb['db_spo coal']
        coal_count = 0
        coal_data = []
        for row in ws_coal.iter_rows(min_row=2, max_row=ws_coal.max_row, values_only=True):
            date_val = row[0]
            pit = row[1]
            losstime_raw = row[2]
            plan_val = row[3]
            
            if not date_val or not pit or not losstime_raw:
                continue
                
            if isinstance(date_val, datetime):
                dt_str = date_val.strftime('%Y-%m-%d')
            else:
                dt_str = str(date_val).split(' ')[0]
                
            try:
                plan_hours = float(plan_val) if plan_val is not None else 0.0
            except ValueError:
                plan_hours = 0.0
                
            # Map losstime
            losstime_raw = str(losstime_raw).strip()
            mapped_name = coal_plan_map.get(losstime_raw, losstime_raw)
            
            coal_data.append((dt_str, pit, mapped_name, plan_hours))
            coal_count += 1
            
        cur.executemany(
            'INSERT INTO plan_spo_coal (date, pit, losstime_name, plan_hours) VALUES (%s, %s, %s, %s)',
            coal_data
        )
        print(f"Inserted {coal_count} rows into plan_spo_coal.")

        # 4. Import OB SPO
        ws_ob = wb['db_spo ob']
        ob_count = 0
        ob_data = []
        for row in ws_ob.iter_rows(min_row=2, max_row=ws_ob.max_row, values_only=True):
            date_val = row[0]
            owner = row[1]
            pit = row[2]
            losstime_raw = row[3]
            plan_val = row[4]
            
            if not date_val or not pit or not losstime_raw:
                continue
                
            if isinstance(date_val, datetime):
                dt_str = date_val.strftime('%Y-%m-%d')
            else:
                dt_str = str(date_val).split(' ')[0]
                
            try:
                plan_hours = float(plan_val) if plan_val is not None else 0.0
            except ValueError:
                plan_hours = 0.0
                
            # Map losstime
            losstime_raw = str(losstime_raw).strip()
            mapped_name = ob_plan_map.get(losstime_raw, losstime_raw)
            
            ob_data.append((dt_str, owner, pit, mapped_name, plan_hours))
            ob_count += 1
            
        cur.executemany(
            'INSERT INTO plan_spo_ob (date, owner, pit, losstime_name, plan_hours) VALUES (%s, %s, %s, %s, %s)',
            ob_data
        )
        print(f"Inserted {ob_count} rows into plan_spo_ob.")
        
        conn.commit()
        print("Transaction committed successfully.")
        
    except Exception as e:
        print(f"ERROR: {e}")
        conn.rollback()
    finally:
        conn.close()

if __name__ == '__main__':
    main()
