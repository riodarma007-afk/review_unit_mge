from fastapi import APIRouter, HTTPException, Depends, File, UploadFile
from typing import List
from app.schemas.settings import TargetSettingCreate, TargetSettingUpdate, TargetSettingResponse
from app.repositories.optrack_repository import OptrackRepository
from app.core.database import get_db_connection
import openpyxl
from io import BytesIO
from datetime import datetime

router = APIRouter()
repo = OptrackRepository()

@router.get("/targets", response_model=List[TargetSettingResponse])
def get_targets(activity: str = None, pit: str = None, year: int = None, month: int = None):
    filters = {}
    if activity: filters['activity'] = activity
    if pit: filters['pit'] = pit
    if year: filters['year'] = year
    if month: filters['month'] = month
    
    records = repo.get_target_settings(**filters)
    return records

@router.post("/targets", response_model=TargetSettingResponse) 
def create_target(setting: TargetSettingCreate):
    setting_data = setting.model_dump()
    new_id = repo.create_target_setting(setting_data)
    
    # Return the newly created record
    records = repo.get_target_settings()
    for rec in records:
        if rec['id'] == new_id:
            return rec
    raise HTTPException(status_code=500, detail="Failed to retrieve created target setting")

@router.put("/targets/{setting_id}", response_model=TargetSettingResponse)
def update_target(setting_id: int, setting: TargetSettingUpdate):
    setting_data = setting.model_dump()
    repo.update_target_setting(setting_id, setting_data)
        
    records = repo.get_target_settings()
    for rec in records:
        if rec['id'] == setting_id:
            return rec
    raise HTTPException(status_code=404, detail="Target setting not found")

@router.delete("/targets/{setting_id}")
def delete_target(setting_id: int):
    deleted = repo.delete_target_setting(setting_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Target setting not found")
    return {"message": "Target setting deleted successfully"}

@router.post("/import-spo")
async def import_spo(file: UploadFile = File(...)):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Hanya file .xlsx yang diperbolehkan")
        
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
        
        # 1. Build Mappings
        ws_map = wb['Standar Penamaan']
        ob_plan_map = {}
        coal_plan_map = {}
        for row in ws_map.iter_rows(min_row=3, max_row=ws_map.max_row, values_only=True):
            if len(row) > 2 and row[2] and row[1]:
                ob_plan_map[str(row[2]).strip()] = str(row[1]).strip()
            if len(row) > 8 and row[8] and row[7]:
                coal_plan_map[str(row[8]).strip()] = str(row[7]).strip()
                
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            # Clear existing data so we can rerun safely
            cur.execute('TRUNCATE TABLE plan_spo_coal')
            cur.execute('TRUNCATE TABLE plan_spo_ob')
            
            # 3. Import Coal SPO
            ws_coal = wb['db_spo coal']
            coal_count = 0
            coal_data = []
            for row in ws_coal.iter_rows(min_row=2, max_row=ws_coal.max_row, values_only=True):
                date_val = row[0]
                pit = row[1]
                losstime_raw = row[2]
                plan_val = row[3]
                
                if not date_val or not pit or not losstime_raw:
                    continue
                    
                if isinstance(date_val, datetime):
                    dt_str = date_val.strftime('%Y-%m-%d')
                else:
                    dt_str = str(date_val).split(' ')[0]
                    
                try:
                    plan_hours = float(plan_val) if plan_val is not None else 0.0
                except ValueError:
                    plan_hours = 0.0
                    
                # Map losstime
                losstime_raw = str(losstime_raw).strip()
                mapped_name = coal_plan_map.get(losstime_raw, losstime_raw)
                
                coal_data.append((dt_str, pit, mapped_name, plan_hours))
                coal_count += 1
                
            cur.executemany(
                'INSERT INTO plan_spo_coal (date, pit, losstime_name, plan_hours) VALUES (%s, %s, %s, %s)',
                coal_data
            )
    
            # 4. Import OB SPO
            ws_ob = wb['db_spo ob']
            ob_count = 0
            ob_data = []
            for row in ws_ob.iter_rows(min_row=2, max_row=ws_ob.max_row, values_only=True):
                date_val = row[0]
                owner = row[1]
                pit = row[2]
                losstime_raw = row[3]
                plan_val = row[4]
                
                if not date_val or not pit or not losstime_raw:
                    continue
                    
                if isinstance(date_val, datetime):
                    dt_str = date_val.strftime('%Y-%m-%d')
                else:
                    dt_str = str(date_val).split(' ')[0]
                    
                try:
                    plan_hours = float(plan_val) if plan_val is not None else 0.0
                except ValueError:
                    plan_hours = 0.0
                    
                # Map losstime
                losstime_raw = str(losstime_raw).strip()
                mapped_name = ob_plan_map.get(losstime_raw, losstime_raw)
                
                ob_data.append((dt_str, owner, pit, mapped_name, plan_hours))
                ob_count += 1
                
            cur.executemany(
                'INSERT INTO plan_spo_ob (date, owner, pit, losstime_name, plan_hours) VALUES (%s, %s, %s, %s, %s)',
                ob_data
            )
            
            conn.commit()
            
            return {
                "message": "File berhasil diproses",
                "coal_rows_inserted": coal_count,
                "ob_rows_inserted": ob_count
            }
            
        except Exception as e:
            conn.rollback()
            raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
        finally:
            conn.close()
            
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Gagal memproses file: {str(e)}")
